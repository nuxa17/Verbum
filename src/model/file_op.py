"""File operations.

This module contains multiple methods for file handling.
"""
import os
import json

import docx
from pdfminer.high_level import extract_text
from striprtf.striprtf import rtf_to_text

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Alignment, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.drawing.image import Image

from model.tools.string_block import StringBlock, StringBlockRegex
from utils.color import ColorGenerator


def read_doc_text(filepath):
    # type: (str) -> str
    """
    Return text from a file based on its extension.
    Supported extensions: `txt`, `doc/docx`, `pdf` and `rtf`.

    Raises:
        `AttributeError`: File not compatible.
    """

    text = ''
    _, extension = os.path.splitext(filepath.lower())

    match extension:
        case ".docx"|".doc":
            doc = docx.Document(filepath)
            for par in doc.paragraphs:
                text += par.text +'\n\n'
        case ".txt":
            with open(filepath, 'r', encoding='UTF-8') as file:
                text = file.read()
        case ".pdf":
            text = extract_text(filepath)
        case ".rtf":
            with open(filepath, 'r', encoding='UTF-8') as file:
                rtf = file.read()
                text = rtf_to_text(rtf)
        case _:
            raise AttributeError("File not compatible.")

    return text


def load_json(filepath):
    # type: (str) -> dict
    """
    Load dictionary from a JSON file.

    Raises:
        `FileNotFoundError`: Could not find the file.
    """
    try:
        with open(filepath, 'r', encoding='UTF-8') as file:
            datos = json.load(file)
    except FileNotFoundError as exc:
        raise FileNotFoundError(f"{filepath} not found.") from exc
    return datos


def save_json(filepath, dictionary):
    """Save dictionary on a JSON file."""
    with open(filepath, 'w', encoding='UTF-8') as file:
        json.dump(dictionary, file)


def save_results(
        filepath, patterns, tags, blob, settings,
        search_results, criteria_results, ngrams_results, graphs_filename=None
    ):
    """
    Save results of the analysis. All the parameters must come from a Model object.
    """
    workbook = Workbook()
    header = NamedStyle(
        name="Header",
        font=Font(bold=True),
        border=Border(bottom=Side(border_style="thin")),
        alignment=Alignment(horizontal="center", vertical="center")
    )
    workbook.add_named_style(header)

    left_border = NamedStyle(
        name="LeftBorder",
        border=Border(
            left=Side(border_style="thin")
        )
    )
    workbook.add_named_style(left_border)

    topleft_border = NamedStyle(
        name="TopLeftBorder",
        border=Border(
            top=Side(border_style="thin"),
            left=Side(border_style="thin")
        )
    )
    workbook.add_named_style(topleft_border)

    sentences = blob.raw_sentences
    patterns_words = [word for category in patterns.values() for word in category["words"]]

    _search_results(workbook, search_results, patterns_words, sentences, settings, tags)
    _criteria_results(workbook, criteria_results)

    if ngrams_results:
        _ngrams_results(workbook, ngrams_results)
    if graphs_filename:
        _sentiment_results(workbook, graphs_filename)

    workbook.save(filename=filepath)

def _search_results(wb, results, patterns_words, sentences, settings, tags):
    # type: (Workbook,list[StringBlock],list[StringBlock],list[str],dict,dict[str, dict]) -> None
    if settings.decoded_tags:
        f_tag = lambda t: tags[t]["tag"]
    else:
        f_tag = lambda t: t

    sheet = wb.active
    sheet.title = "Patterns"

    cols = {"Category": 1, "String": 2, "Tags": 3,
            "Text posit.": 4, "Sent. num.": 5, "Sentence": 6}

    #**********
    #* HEADER *
    #**********

    for c in cols.items():
        sheet.cell(1, c[1]).value = c[0]
        sheet.cell(1, c[1]).style = "Header"

    sheet.freeze_panes = "C2"  # fix category, string and header

    #**********
    #* CELLS *
    #**********

    nrow = 2
    last_cat = None

    for idx, result in enumerate(results):    # fill patterns
        res = result
        pat = patterns_words[idx]

        if isinstance(pat, StringBlockRegex):
            name = pat["meaning"] + "("+ pat["string"] + ")"
            tot_pos = len([elem for sublist in res.pos_text for elem in sublist])
        else:
            name = pat["string"]
            tot_pos = len(res.pos_text)

        if tot_pos == 0 and not settings.unmatched:
            continue

        # category changed
        if last_cat is None or res.category != last_cat:
            last_cat = res.category
            sheet.cell(nrow, cols['Category']).value = res.category
            sheet.cell(nrow, cols['Category']).style = "TopLeftBorder"

        for c in cols.items():  # first row
            if c[0] != 'Category': # avoid category
                sheet.cell(nrow, c[1]).style = "TopLeftBorder"

        sheet.cell(nrow, cols['String']).value = name
        sheet.cell(nrow + 1, cols['String']).value = "Total: " + str(tot_pos)

        count_row = nrow    # continue on the same row
        if isinstance(res, StringBlockRegex):
            for num_sb, _ in enumerate(res.reg_matched):
                matched = res.reg_matched[num_sb]
                words = matched if matched != "" else res.words
                aux_sb = StringBlock(
                    words, pos_text=res.pos_text[num_sb],
                    n_sent=res.n_sent[num_sb], pos_sent=res.pos_sent[num_sb],
                    tags=res.tags[num_sb], category=res.category
                )
                count_row = _search_results_aux(
                    sheet, cols, sentences, settings, tags, pat, aux_sb, count_row
                )
        else:
            count_row =_search_results_aux(
                sheet, cols, sentences, settings, tags, pat, res, count_row
            )

        if (tot_pos == 0 and settings.save_tags and "tags" in pat):
            this_tags = [f_tag(t) for t in pat["tags"]]
            sheet.cell(nrow, cols['Tags']).value = '; '.join(this_tags)

        tot_pos = max(tot_pos, 2)

        for i in range(1, tot_pos):
            for c in cols.items():  # first row
                sheet.cell(nrow + i, c[1]).style = "LeftBorder"

        nrow += tot_pos

    # columns width
    sheet.column_dimensions[get_column_letter(cols["Category"])].width = 25
    sheet.column_dimensions[get_column_letter(cols["String"])].width = 20
    if settings.save_tags:
        sheet.column_dimensions[get_column_letter(cols["Tags"])].width = 20
    else: sheet.delete_cols(cols["Tags"])

def _search_results_aux(sheet, cols, sentences, settings, tags, pat, res:StringBlock, nrow):
    highlight = InlineFont(color=settings.last_color[1:], b=True)
    letter_frase = get_column_letter(cols['Sentence'])

    if settings.decoded_tags:
        f_tag = lambda t: tags[t]["tag"]
    else:
        f_tag = lambda t: t

    ocurr=0
    for ocurr, _ in enumerate(res.pos_text):
        sheet.cell(nrow + ocurr, cols['Text posit.']).value = res.pos_text[ocurr]
        sheet.cell(nrow + ocurr, cols['Sent. num.']).value = res.n_sent[ocurr]

        sent = sentences[res.n_sent[ocurr]].split()
        posit = list(set(res.pos_sent[ocurr]))    #remove duplicates from contractions

        sent_text = CellRichText()
        if posit[0] > 0:
            sent_text.append(" ".join(sent[:posit[0]]) + " ")
        for pos in posit:
            sent_text.append(TextBlock(highlight, sent[pos]+ " "))
        if posit[-1] < len(sent)-1:
            sent_text.append(" ".join(sent[posit[-1]+1:]))

        sheet[letter_frase +str(nrow + ocurr)] = sent_text

        if not settings.save_tags:  # don't save them
            pass
        elif "tags" in pat: # both tags
            val = []

            for r_tag, p_tag in zip(res.tags[ocurr], pat["tags"]):
                if r_tag != p_tag: # match
                    val.append(f_tag(r_tag) + '(' + f_tag(p_tag) + ')')
                else:
                    val.append(f_tag(r_tag))

            sheet.cell(nrow + ocurr, cols['Tags']).value = '; '.join(val)
        else:   # result only
            this_tags = [f_tag(t) for t in res.tags[ocurr]]
            sheet.cell(nrow + ocurr, cols['Tags']).value = '; '.join(this_tags)

    return nrow + ocurr +1  # ocurr begins with 0

def _criteria_results(wb, results):
    sheet = wb.create_sheet("Manipulation rates")

    cols = {"Category": 1, "Criteria": 2, "Result": 3,
            "Found": 4, "Against": 5, "Percentage": 6}

    #**********
    #* HEADER *
    #**********

    for c in cols.items():
        sheet.cell(1, c[1]).value = c[0]
        sheet.cell(1, c[1]).style = "Header"

    sheet.freeze_panes = "C2"  # fix category, string and header

    cols = {"category": 1, "criteria": 2, "str": 3,
            "found": 4, "against": 5, "percentage": 6}

    colors = ColorGenerator.get_gradient()
    nrow =2
    for category, info in results.items():
        sheet.cell(nrow, cols["category"]).value = category
        sheet.cell(nrow, 1).style = "TopLeftBorder"

        rank = info.pop("rank")
        color = colors[rank]
        foreground = ColorGenerator.best_foreground(color)
        word_color = InlineFont(color=foreground[1:])

        for idx, (k, v) in enumerate(info.items()):
            if k == "str":
                sheet.cell(nrow, cols[k]).value = CellRichText(TextBlock(word_color, v))
            else:
                sheet.cell(nrow, cols[k]).value = v
            sheet.cell(nrow, idx+2).style = "TopLeftBorder"

        fill_color = Color(color[1:])
        fill = PatternFill(start_color=fill_color,
                   end_color=fill_color,
                   fill_type='solid')
        sheet.cell(nrow, cols["str"]).fill = fill

        nrow += 1

    sheet.column_dimensions[get_column_letter(cols["category"])].width = 30
    sheet.column_dimensions[get_column_letter(cols["criteria"])].width = 40
    sheet.column_dimensions[get_column_letter(cols["str"])].width = 20

def _ngrams_results(wb, results):
    # type: (Workbook, list[tuple[dict, tuple]]) -> None
    for num, (query, ngrams) in enumerate(results):
        sheet = wb.create_sheet("N-grams " + str(num+1))

        #**********
        #* HEADER *
        #**********

        for i, (k, v) in enumerate(query.items()):
            i += 8
            sheet.cell(1, i).value = k
            sheet.cell(1, i).style = "Header"
            sheet.cell(2, i).value = v if not isinstance(v,(list, tuple)) else ', '.join(v)
            sheet.cell(2, i).style = "TopLeftBorder"
            if k in ["Removed", "Contains"]:
                sheet.column_dimensions[get_column_letter(i)].width = 20
            else:
                sheet.column_dimensions[get_column_letter(i)].width = 10


        cols = {"N-gram":1, "Frecuency": 2, "Sentences":3}

        for c in cols.items():
            sheet.cell(3, c[1]).value = c[0]
            sheet.cell(3, c[1]).style = "Header"

        sheet.freeze_panes = "C4"  # fix n-gram

        #**********
        #* CELLS *
        #**********

        nrow = 4

        for words, sentences in ngrams:
            sheet.cell(nrow, cols['N-gram']).value = "(" + ", ".join(words) + ")"
            sheet.cell(nrow, cols['Frecuency']).value = len(sentences)

            for c in cols.items():  # first row
                sheet.cell(nrow, c[1]).style = "TopLeftBorder"

            # one occurrence per row
            ocurr=0
            for ocurr, sent in enumerate(sentences):
                sheet.cell(nrow + ocurr, cols['Sentences']).value = sent
                # style
                if ocurr == 0:  # skip the first row
                    continue
                for c in cols.items():
                    sheet.cell(nrow + ocurr, c[1]).style = "LeftBorder"

            nrow += ocurr+1

        sheet.column_dimensions[get_column_letter(cols["N-gram"])].width = 20
        sheet.column_dimensions[get_column_letter(cols["Frecuency"])].width = 10
        sheet.column_dimensions[get_column_letter(cols["Sentences"])].width = 10

def _sentiment_results(wb, graphs_filename):
    sheet = wb.create_sheet("Sentiment")
    img = Image(graphs_filename)
    img.anchor = 'A1'
    sheet.add_image(img)
